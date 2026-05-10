import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def export_to_excel(test_cases, product, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{product} Test Cases"

    # Styles
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(fill_type='solid', fgColor='534AB7')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        'Serial No.', 'Date', 'Test Suite ID', 'Test Case ID',
        'Test Case Objective', 'Preconditions', 'Test Steps',
        'Test Data', 'Expected Result', 'Actual Result', 'Status', 'Remarks'
    ]

    col_widths = [10, 12, 15, 15, 40, 30, 50, 25, 40, 30, 12, 20]

    # Write headers
    for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_num)
        ].width = width

    ws.row_dimensions[1].height = 30

    # Write test cases
    for row_num, tc in enumerate(test_cases, 2):
        values = [
            tc.get('serial', row_num - 1),
            tc.get('date', datetime.now().strftime('%d/%m/%Y')),
            tc.get('suite_id', ''),
            tc.get('tc_id', ''),
            tc.get('objective', ''),
            tc.get('preconditions', ''),
            tc.get('steps', ''),
            tc.get('test_data', ''),
            tc.get('expected_result', ''),
            tc.get('actual_result', ''),
            tc.get('status', ''),
            tc.get('remarks', '')
        ]

        row_color = 'F8F8FC' if row_num % 2 == 0 else 'FFFFFF'

        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = Font(name='Arial', size=9)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.fill = PatternFill(fill_type='solid', fgColor=row_color)
            cell.border = border

        ws.row_dimensions[row_num].height = 60

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    return output_path